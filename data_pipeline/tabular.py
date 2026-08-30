from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from data_pipeline.delimited import validate_delimited_shape
from data_pipeline.validation import (
    IDENTITY_MAX_LENGTHS,
    DataValidationError,
    SalesFrameValidator,
)

SUPPORTED_EXTENSIONS = {".csv": "csv", ".tsv": "tsv", ".xlsx": "xlsx"}
MAX_TABULAR_ROWS = 500_000
MAX_TABULAR_COLUMNS = 100
MAX_TABULAR_CELLS = 5_000_000
MAX_CELL_CHARACTERS = 10_000
MAX_XLSX_MEMBERS = 1_000
MAX_XLSX_EXPANDED_BYTES = 128 * 1024 * 1024
MAX_XLSX_MEMBER_BYTES = 64 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 100
MAX_WORKSHEETS = 50
MAX_HEADER_CHARACTERS = 255
PREVIEW_ROWS = 5
PREVIEW_VALUE_CHARACTERS = 160

MAPPING_FIELDS = (
    "order_id",
    "order_date",
    "customer_id",
    "region",
    "category",
    "product",
    "quantity",
    "unit_price",
    "discount",
    "revenue",
    "currency",
)

FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "order_id": (
        "order_id",
        "order_number",
        "order_no",
        "invoice",
        "invoice_id",
        "invoice_no",
        "transaction_id",
        "receipt_id",
        "sale_id",
        "订单号",
        "訂單號",
        "交易号",
        "交易號",
    ),
    "order_date": (
        "order_date",
        "date",
        "sales_date",
        "sale_date",
        "transaction_date",
        "invoice_date",
        "ordered_on",
        "timestamp",
        "日期",
        "销售日期",
        "銷售日期",
        "订单日期",
        "訂單日期",
        "交易日期",
    ),
    "customer_id": (
        "customer_id",
        "customer",
        "customer_number",
        "client_id",
        "buyer_id",
        "account_id",
        "store_id",
        "store_no",
        "客户",
        "客戶",
        "客户编号",
        "客戶編號",
        "门店",
        "門店",
    ),
    "region": (
        "region",
        "country",
        "state",
        "province",
        "county",
        "county_name",
        "territory",
        "market",
        "location",
        "地区",
        "地區",
        "区域",
        "區域",
    ),
    "category": (
        "category",
        "category_name",
        "department",
        "department_name",
        "dept",
        "segment",
        "product_category",
        "类别",
        "類別",
        "品类",
        "品類",
    ),
    "product": (
        "product",
        "product_name",
        "item",
        "item_name",
        "description",
        "sku",
        "stock_code",
        "产品",
        "產品",
        "商品",
    ),
    "quantity": (
        "quantity",
        "qty",
        "units",
        "units_sold",
        "sales_units",
        "sales_bottles",
        "demand",
        "数量",
        "數量",
        "销量",
        "銷量",
    ),
    "unit_price": (
        "unit_price",
        "price",
        "sell_price",
        "sales_price",
        "item_price",
        "price_each",
        "单价",
        "單價",
    ),
    "discount": (
        "discount",
        "discount_rate",
        "discount_pct",
        "discount_percent",
        "折扣",
    ),
    "revenue": (
        "revenue",
        "sales",
        "net_sales",
        "sales_amount",
        "sales_dollars",
        "amount",
        "total",
        "line_total",
        "net_amount",
        "销售额",
        "銷售額",
        "营收",
        "營收",
        "金额",
        "金額",
    ),
    "currency": (
        "currency",
        "currency_code",
        "currency_code_value",
        "currency_cd",
        "currency_description",
        "currency_desc",
        "currency_type",
        "currency_name",
        "iso_currency",
        "currency_iso",
        "transaction_currency",
        "source_currency",
        "local_currency",
        "base_currency",
        "reporting_currency",
        "document_currency",
        "payment_currency",
        "settlement_currency",
        "invoice_currency",
        "order_currency",
        "ccy",
        "ccy_code",
        "curr",
        "币种",
        "幣種",
        "货币",
        "貨幣",
        "交易币种",
        "交易幣種",
        "来源币种",
        "來源幣種",
    ),
}

FIELD_DEFINITIONS = (
    {
        "name": "order_date",
        "label": "Order date",
        "required": True,
        "description": "Calendar date used for trends and forecasting.",
        "aliases": list(FIELD_ALIASES["order_date"]),
        "default": None,
    },
    {
        "name": "revenue",
        "label": "Revenue",
        "required": False,
        "description": "Direct non-negative net revenue; use instead of price components.",
        "aliases": list(FIELD_ALIASES["revenue"]),
        "default": None,
    },
    {
        "name": "quantity",
        "label": "Quantity",
        "required": False,
        "description": "Required with unit price unless direct revenue is mapped.",
        "aliases": list(FIELD_ALIASES["quantity"]),
        "default": "1 when direct revenue is used",
    },
    {
        "name": "unit_price",
        "label": "Unit price",
        "required": False,
        "description": "Required with quantity unless direct revenue is mapped.",
        "aliases": list(FIELD_ALIASES["unit_price"]),
        "default": None,
    },
    {
        "name": "discount",
        "label": "Discount",
        "required": False,
        "description": "Fraction or percentage applied to component-based revenue.",
        "aliases": list(FIELD_ALIASES["discount"]),
        "default": 0,
    },
    {
        "name": "currency",
        "label": "Currency code",
        "required": False,
        "description": (
            "Optional per-row USD or GBP code; required when the file has an "
            "obvious currency column."
        ),
        "aliases": list(FIELD_ALIASES["currency"]),
        "default": "Selected source currency",
    },
    {
        "name": "order_id",
        "label": "Order ID",
        "required": False,
        "description": "Blank, missing, or repeated IDs are deterministically disambiguated.",
        "aliases": list(FIELD_ALIASES["order_id"]),
        "default": "Generated sales-record ID",
    },
    {
        "name": "customer_id",
        "label": "Customer or entity ID",
        "required": False,
        "description": "Required for meaningful entity segmentation.",
        "aliases": list(FIELD_ALIASES["customer_id"]),
        "default": "UNSPECIFIED-ENTITY",
    },
    {
        "name": "region",
        "label": "Region",
        "required": False,
        "description": "Geographic or market grouping.",
        "aliases": list(FIELD_ALIASES["region"]),
        "default": "All regions",
    },
    {
        "name": "category",
        "label": "Category",
        "required": False,
        "description": "Product or service category.",
        "aliases": list(FIELD_ALIASES["category"]),
        "default": "Uncategorized",
    },
    {
        "name": "product",
        "label": "Product",
        "required": False,
        "description": "Product, item, SKU, or service name.",
        "aliases": list(FIELD_ALIASES["product"]),
        "default": "Unspecified product",
    },
)


@dataclass(frozen=True)
class TabularDataset:
    frame: pd.DataFrame
    filename: str
    file_format: str
    file_sha256: str
    sheets: list[str]
    selected_sheet: str | None


@dataclass(frozen=True)
class MappedSalesDataset:
    frame: pd.DataFrame
    mapping: dict[str, str]
    metric_mode: str
    generated_fields: list[str]
    warnings: list[str]
    record_semantics: dict[str, Any]


def _canonical_header(value: object) -> str:
    return SalesFrameValidator.canonical_name(value)


def _header_match_key(value: object) -> str:
    normalized = unicodedata.normalize("NFKC", str(value).strip()).casefold()
    return re.sub(r"[\W_]+", "", normalized, flags=re.UNICODE)


def _validate_headers(headers: list[object]) -> list[str]:
    names = ["" if value is None else str(value).strip() for value in headers]
    blank_count = sum(not name for name in names)
    if blank_count:
        raise DataValidationError(
            [f"The table header has {blank_count} blank column names."]
        )
    oversized = [name for name in names if len(name) > MAX_HEADER_CHARACTERS]
    if oversized:
        raise DataValidationError(
            [
                "The table header has "
                f"{len(oversized)} column names longer than "
                f"{MAX_HEADER_CHARACTERS} characters."
            ]
        )
    canonical = [_header_match_key(name) for name in names]
    if any(not value for value in canonical):
        raise DataValidationError(
            ["Column names must contain at least one letter or number."]
        )
    duplicates = sorted(
        name for name, count in Counter(canonical).items() if count > 1
    )
    if duplicates:
        raise DataValidationError(
            [
                "Column names collide after normalization: "
                f"{', '.join(duplicates)}."
            ]
        )
    return names


def _validate_frame_limits(frame: pd.DataFrame) -> None:
    issues: list[str] = []
    rows, columns = frame.shape
    if rows == 0:
        issues.append("The uploaded table contains no data rows.")
    if rows > MAX_TABULAR_ROWS:
        issues.append(f"The table exceeds the {MAX_TABULAR_ROWS:,}-row limit.")
    if columns > MAX_TABULAR_COLUMNS:
        issues.append(
            f"The table exceeds the {MAX_TABULAR_COLUMNS:,}-column limit."
        )
    if rows * columns > MAX_TABULAR_CELLS:
        issues.append(
            f"The table exceeds the {MAX_TABULAR_CELLS:,}-cell processing limit."
        )
    for column in frame.columns:
        values = frame[column].astype("string")
        nul_count = int(values.str.contains("\x00", regex=False, na=False).sum())
        if nul_count:
            issues.append(
                f"Column '{column}' has {nul_count} values containing a NUL "
                "control character."
            )
        too_long = values.str.len().gt(MAX_CELL_CHARACTERS).fillna(False)
        if too_long.any():
            issues.append(
                f"Column '{column}' has {int(too_long.sum())} values longer than "
                f"{MAX_CELL_CHARACTERS:,} characters."
            )
        if len(issues) >= 20:
            break
    if issues:
        raise DataValidationError(issues)


def _text_frame(content: bytes, *, delimiter: str | None = None) -> pd.DataFrame:
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV and TSV files must use UTF-8 encoding.") from exc
    if "\x00" in decoded:
        raise DataValidationError(
            ["The uploaded table contains a NUL control character."]
        )
    if not decoded.strip():
        raise ValueError("Unable to parse table: the file is empty.")
    if delimiter is None:
        try:
            delimiter = csv.Sniffer().sniff(
                decoded[:64 * 1024], delimiters=",;\t|"
            ).delimiter
        except csv.Error:
            delimiter = ","
    validate_delimited_shape(
        decoded,
        delimiter,
        max_columns=MAX_TABULAR_COLUMNS,
        max_header_characters=MAX_HEADER_CHARACTERS,
        max_cell_characters=MAX_CELL_CHARACTERS,
    )
    try:
        reader = csv.reader(
            StringIO(decoded, newline=""), delimiter=delimiter, strict=True
        )
        header = next(reader)
        if len(header) > MAX_TABULAR_COLUMNS:
            raise DataValidationError(
                [f"The table exceeds the {MAX_TABULAR_COLUMNS:,}-column limit."]
            )
        names = _validate_headers(header)
        rows: list[list[str]] = []
        for data_row_number, row in enumerate(reader, start=1):
            if not row or not any(value.strip() for value in row):
                continue
            if len(row) != len(names):
                raise DataValidationError(
                    [
                        f"Data row {data_row_number} has {len(row)} fields; "
                        f"the header defines {len(names)}."
                    ]
                )
            rows.append(row)
            if len(rows) > MAX_TABULAR_ROWS:
                raise DataValidationError(
                    [f"The table exceeds the {MAX_TABULAR_ROWS:,}-row limit."]
                )
            if len(rows) * len(names) > MAX_TABULAR_CELLS:
                raise DataValidationError(
                    [
                        "The table exceeds the "
                        f"{MAX_TABULAR_CELLS:,}-cell processing limit."
                    ]
                )
        frame = pd.DataFrame(rows, columns=names, dtype="string")
    except StopIteration as exc:
        raise ValueError("Unable to parse table: the file is empty.") from exc
    except (csv.Error, pd.errors.ParserError) as exc:
        raise ValueError(f"Unable to parse table: {exc}") from exc
    frame.columns = names
    return frame


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_MEMBERS:
                raise ValueError(
                    f"XLSX archive exceeds the {MAX_XLSX_MEMBERS:,}-member limit."
                )
            expanded = sum(member.file_size for member in members)
            if expanded > MAX_XLSX_EXPANDED_BYTES:
                raise ValueError(
                    "XLSX expanded content exceeds the "
                    f"{MAX_XLSX_EXPANDED_BYTES // (1024 * 1024)} MiB limit."
                )
            for member in members:
                normalized = member.filename.replace("\\", "/").lower()
                if member.flag_bits & 0x1:
                    raise ValueError("Encrypted XLSX archives are not supported.")
                if member.file_size > MAX_XLSX_MEMBER_BYTES:
                    raise ValueError(
                        "An XLSX archive member exceeds the safe expanded-size limit."
                    )
                ratio = member.file_size / max(member.compress_size, 1)
                if ratio > MAX_XLSX_COMPRESSION_RATIO:
                    raise ValueError(
                        "XLSX compression ratio exceeds the safe processing limit."
                    )
                if normalized.endswith("vbaproject.bin"):
                    raise ValueError("Macro-enabled workbooks are not supported.")
                if normalized.startswith("xl/externallinks/"):
                    raise ValueError("Workbooks with external links are not supported.")
    except zipfile.BadZipFile as exc:
        raise ValueError("Unable to parse XLSX: the file is not a valid workbook.") from exc


def _xlsx_frame(
    content: bytes, sheet_name: str | None
) -> tuple[pd.DataFrame, list[str], str]:
    _validate_xlsx_archive(content)
    try:
        workbook = load_workbook(
            BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        raise ValueError(f"Unable to parse XLSX workbook: {exc}") from exc
    try:
        sheets = list(workbook.sheetnames)
        if not sheets:
            raise ValueError("The XLSX workbook contains no worksheets.")
        if len(sheets) > MAX_WORKSHEETS:
            raise ValueError(
                f"The XLSX workbook exceeds the {MAX_WORKSHEETS}-worksheet limit."
            )
        selected = sheet_name or sheets[0]
        if selected not in sheets:
            raise ValueError(f"Worksheet '{selected}' does not exist in the workbook.")
        worksheet = workbook[selected]
        if worksheet.max_column and worksheet.max_column > MAX_TABULAR_COLUMNS:
            raise DataValidationError(
                [f"The table exceeds the {MAX_TABULAR_COLUMNS:,}-column limit."]
            )
        if worksheet.max_row and worksheet.max_row > MAX_TABULAR_ROWS + 1:
            raise DataValidationError(
                [f"The table exceeds the {MAX_TABULAR_ROWS:,}-row limit."]
            )
        if (
            worksheet.max_column
            and worksheet.max_row
            and worksheet.max_column * max(0, worksheet.max_row - 1)
            > MAX_TABULAR_CELLS
        ):
            raise DataValidationError(
                [
                    "The table exceeds the "
                    f"{MAX_TABULAR_CELLS:,}-cell processing limit."
                ]
            )
        reset_dimensions = getattr(worksheet, "reset_dimensions", None)
        if callable(reset_dimensions):
            reset_dimensions()
        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_header = list(next(rows))
        except StopIteration as exc:
            raise ValueError("The selected worksheet is empty.") from exc
        while raw_header and raw_header[-1] is None:
            raw_header.pop()
        if len(raw_header) > MAX_TABULAR_COLUMNS:
            raise DataValidationError(
                [f"The table exceeds the {MAX_TABULAR_COLUMNS:,}-column limit."]
            )
        headers = _validate_headers(raw_header)
        values: list[list[object]] = []
        for physical_data_row, row in enumerate(rows, start=1):
            if physical_data_row > MAX_TABULAR_ROWS:
                raise DataValidationError(
                    [f"The table exceeds the {MAX_TABULAR_ROWS:,}-row limit."]
                )
            if physical_data_row * len(headers) > MAX_TABULAR_CELLS:
                raise DataValidationError(
                    [
                        "The table exceeds the "
                        f"{MAX_TABULAR_CELLS:,}-cell processing limit."
                    ]
                )
            row_values = list(row)
            extras = row_values[len(headers) :]
            if any(value is not None and str(value).strip() for value in extras):
                raise DataValidationError(
                    [
                        f"Worksheet data row {physical_data_row} contains values "
                        "beyond the last named header column."
                    ]
                )
            normalized = row_values[: len(headers)]
            if not any(value is not None and str(value).strip() for value in normalized):
                continue
            values.append(normalized)
        return pd.DataFrame(values, columns=headers, dtype=object), sheets, selected
    except (DataValidationError, ValueError):
        raise
    except Exception as exc:
        raise ValueError("Unable to parse XLSX workbook content.") from exc
    finally:
        workbook.close()


def read_tabular_file(
    content: bytes, filename: str, *, sheet_name: str | None = None
) -> TabularDataset:
    safe_filename = re.split(r"[/\\\\]", filename)[-1]
    if not safe_filename or len(safe_filename) > 255 or "\x00" in safe_filename:
        raise ValueError("Uploaded filename is invalid or exceeds 255 characters.")
    extension = Path(safe_filename).suffix.lower()
    try:
        file_format = SUPPORTED_EXTENSIONS[extension]
    except KeyError as exc:
        raise ValueError("Upload must be a .csv, .tsv, or .xlsx file.") from exc
    digest = hashlib.sha256(content).hexdigest()
    if file_format == "xlsx":
        if sheet_name is not None and (
            not sheet_name or len(sheet_name) > 31 or "\x00" in sheet_name
        ):
            raise ValueError("sheet_name is invalid for an XLSX workbook.")
        frame, sheets, selected_sheet = _xlsx_frame(content, sheet_name)
    else:
        if sheet_name:
            raise ValueError("sheet_name is only valid for XLSX workbooks.")
        delimiter = "\t" if file_format == "tsv" else None
        frame = _text_frame(content, delimiter=delimiter)
        sheets = []
        selected_sheet = None
    _validate_frame_limits(frame)
    return TabularDataset(
        frame=frame,
        filename=safe_filename,
        file_format=file_format,
        file_sha256=digest,
        sheets=sheets,
        selected_sheet=selected_sheet,
    )


_NUMERIC_CURRENCY_TOKEN = (
    r"(?:USD|GBP|EUR|CAD|AUD|CNY|RMB|JPY|INR|US\$|[$£€¥₹])"
)
_NUMERIC_BODY = (
    r"(?:"
    r"(?:\d{1,3}(?:,\d{3})+)(?:\.\d+)?|"
    r"(?:\d{1,3}(?:\.\d{3})+)(?:,\d+)?|"
    r"(?:\d{1,3}(?:[ '\u00a0]\d{3})+)(?:[.,]\d+)?|"
    r"(?:\d+(?:[.,]\d+)?|\.\d+)(?:[eE][+-]?\d+)?"
    r")"
)
_STRICT_NUMERIC_PATTERN = re.compile(
    rf"^(?P<outer_sign>[+-])?\s*"
    rf"(?P<prefix>{_NUMERIC_CURRENCY_TOKEN})?\s*"
    rf"(?P<inner_sign>[+-])?\s*"
    rf"(?P<number>{_NUMERIC_BODY})\s*"
    rf"(?P<suffix>{_NUMERIC_CURRENCY_TOKEN})?$",
    re.IGNORECASE,
)


def _normalize_numeric_text(value: object) -> str | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        return str(value)
    text = str(value).strip()
    if not text:
        return None
    parenthesized = text.startswith("(") and text.endswith(")")
    if parenthesized:
        text = text[1:-1]
    elif "(" in text or ")" in text:
        return None
    match = _STRICT_NUMERIC_PATTERN.fullmatch(text.strip())
    if match is None:
        return None
    if match.group("prefix") and match.group("suffix"):
        return None
    if match.group("outer_sign") and match.group("inner_sign"):
        return None
    if parenthesized and (match.group("outer_sign") or match.group("inner_sign")):
        return None
    number = match.group("number").replace(" ", "").replace("\u00a0", "")
    number = number.replace("'", "")
    if "," in number and "." in number:
        if number.rfind(",") > number.rfind("."):
            number = number.replace(".", "").replace(",", ".")
        else:
            number = number.replace(",", "")
    elif "," in number:
        if re.fullmatch(r"\d{1,3}(?:,\d{3})+", number):
            number = number.replace(",", "")
        else:
            number = number.replace(",", ".")
    elif number.count(".") > 1:
        number = number.replace(".", "")
    sign = match.group("outer_sign") or match.group("inner_sign") or ""
    if parenthesized:
        sign = "-"
    return f"{sign}{number}"


_CURRENCY_CODE_PATTERN = re.compile(
    r"(?i)(?<![A-Z])(?:USD|GBP|EUR|CAD|AUD|CNY|RMB|JPY|INR|CHF|NZD|SGD|"
    r"HKD|KRW|MXN|BRL|ZAR|SEK|NOK|DKK)(?![A-Z])"
)
_CURRENCY_SYMBOLS = {
    "$": "USD",
    "£": "GBP",
    "€": "EUR",
    "¥": "JPY",
    "₹": "INR",
}
_KNOWN_CURRENCY_CODES = {
    "USD",
    "GBP",
    "EUR",
    "CAD",
    "AUD",
    "CNY",
    "RMB",
    "JPY",
    "INR",
    "CHF",
    "NZD",
    "SGD",
    "HKD",
    "KRW",
    "MXN",
    "BRL",
    "ZAR",
    "SEK",
    "NOK",
    "DKK",
}


def _currency_cell_code(value: object) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip().upper()
    symbol_values = {"$": "USD", "US$": "USD", "£": "GBP", "€": "EUR", "¥": "JPY", "₹": "INR"}
    if text in symbol_values:
        return symbol_values[text]
    if text in _KNOWN_CURRENCY_CODES:
        return "CNY" if text == "RMB" else text
    return None


def _strong_currency_header(value: object) -> bool:
    normalized = unicodedata.normalize("NFKC", str(value).strip()).casefold()
    tokens = re.findall(r"[^\W_]+", normalized, flags=re.UNICODE)
    key = _header_match_key(value)
    currency_tokens = {
        "currency",
        "ccy",
        "币种",
        "幣種",
        "货币",
        "貨幣",
    }
    qualifier_tokens = {"code", "description", "desc", "type", "iso", "name"}
    normalized_codes = {code.casefold() for code in _KNOWN_CURRENCY_CODES}
    has_currency_token = any(token in currency_tokens for token in tokens)
    return (
        key.endswith(("币种", "幣種", "货币", "貨幣"))
        or (
            has_currency_token
            and (
                len(tokens) == 1
                or tokens[-1] in currency_tokens
                or any(token in qualifier_tokens for token in tokens)
                or any(token in normalized_codes for token in tokens)
            )
        )
    )


def _currency_candidate_columns(
    frame: pd.DataFrame, *, include_value_only: bool = True
) -> list[str]:
    alias_keys = {_header_match_key(alias) for alias in FIELD_ALIASES["currency"]}
    candidates: list[str] = []
    for column in frame.columns:
        name = str(column)
        key = _header_match_key(name)
        populated = frame[column].dropna().astype("string").str.strip()
        populated = populated[populated.ne("")]
        exact_header_match = key in alias_keys
        strong_header_match = _strong_currency_header(name)
        broad_header_match = "currency" in key or any(
            token in key for token in ("币种", "幣種", "货币", "貨幣")
        )
        recognized = populated.map(
            lambda value: _currency_cell_code(value) is not None
        )
        any_value_match = bool(len(populated)) and recognized.any()
        all_values_match = bool(len(populated)) and recognized.all()
        if (
            exact_header_match
            or strong_header_match
            or (broad_header_match and any_value_match)
            or (include_value_only and all_values_match)
        ):
            candidates.append(name)
    return candidates


def _currency_markers(series: pd.Series) -> set[str]:
    markers: set[str] = set()
    for value in series.dropna():
        if isinstance(value, (int, float, np.integer, np.floating)):
            continue
        text = str(value)
        markers.update(
            "CNY" if match.upper() == "RMB" else match.upper()
            for match in _CURRENCY_CODE_PATTERN.findall(text)
        )
        markers.update(
            code for symbol, code in _CURRENCY_SYMBOLS.items() if symbol in text
        )
    return markers


def _validate_monetary_currency(
    series: pd.Series, *, field: str, source_currency: str
) -> None:
    markers = _currency_markers(series)
    markers.update(_currency_markers(pd.Series([series.name], dtype="string")))
    unsupported = sorted(markers.difference({"USD", "GBP"}))
    if unsupported:
        raise DataValidationError(
            [
                f"Mapped '{field}' contains unsupported currency markers: "
                f"{', '.join(unsupported)}. Convert values upstream without mixing "
                "currencies."
            ]
        )
    mismatched = sorted(markers.difference({source_currency}))
    if mismatched:
        raise DataValidationError(
            [
                f"Mapped '{field}' contains {', '.join(mismatched)} values but the "
                f"selected source currency is {source_currency}."
            ]
        )


def _reject_currency_markers(series: pd.Series, *, field: str) -> None:
    if _currency_markers(series):
        raise DataValidationError(
            [f"Mapped '{field}' cannot contain currency symbols or currency codes."]
        )


def _reject_percent_markers(series: pd.Series, *, field: str) -> None:
    raw = series.astype("string")
    marked = raw.str.contains("%", regex=False, na=False)
    if marked.any():
        raise DataValidationError(
            [
                f"Mapped '{field}' contains {int(marked.sum())} percentage values; "
                "percent notation is only valid for the discount field."
            ]
        )


def _validate_currency_column(series: pd.Series, *, source_currency: str) -> None:
    normalized: list[str] = []
    blank = 0
    invalid_codes: set[str] = set()
    for value in series:
        if value is None or pd.isna(value) or not str(value).strip():
            blank += 1
            continue
        resolved = _currency_cell_code(value)
        if resolved is None:
            safe_value = "".join(
                character
                for character in str(value).strip().upper()[:20]
                if ord(character) >= 32 and ord(character) != 127
            )
            invalid_codes.add(safe_value or "<invalid>")
            continue
        normalized.append(resolved)
    issues: list[str] = []
    if blank:
        issues.append(
            f"Mapped 'currency' contains {blank} blank values; every imported row "
            "must declare one source currency."
        )
    if invalid_codes:
        examples = ", ".join(sorted(invalid_codes)[:5])
        issues.append(
            "Mapped 'currency' contains unsupported or invalid codes: "
            f"{examples}. Convert the source to USD or GBP first."
        )
    if issues:
        raise DataValidationError(
            issues
        )
    unsupported = sorted(set(normalized).difference({"USD", "GBP"}))
    if unsupported:
        raise DataValidationError(
            [
                "Mapped 'currency' contains unsupported codes: "
                f"{', '.join(unsupported)}. Convert the source to USD or GBP first."
            ]
        )
    mismatched = sorted(set(normalized).difference({source_currency}))
    if mismatched:
        raise DataValidationError(
            [
                "Mapped 'currency' does not match the selected source currency "
                f"{source_currency}: found {', '.join(mismatched)}."
            ]
        )


def _parse_discount(series: pd.Series, source_name: str) -> tuple[pd.Series, bool]:
    raw = series.astype("string").str.strip()
    populated = raw.notna() & raw.ne("")
    contains_percent = raw.str.contains("%", regex=False, na=False)
    percent_markers = raw.str.fullmatch(r"[^%]+%", na=False)
    invalid_percent = contains_percent & ~percent_markers
    if invalid_percent.any():
        raise DataValidationError(
            [
                "Mapped 'discount' contains malformed percentage notation; use "
                "at most one trailing '%' marker per value."
            ]
        )
    numeric_raw = raw.mask(percent_markers, raw.str[:-1].str.strip())
    parsed = _parse_numbers(numeric_raw)
    header_key = _header_match_key(source_name)
    percent_header = "%" in source_name or any(
        token in header_key for token in ("pct", "percent", "percentage", "百分比")
    )
    if percent_markers.any() and not percent_header:
        unmarked = populated & ~percent_markers
        if unmarked.any():
            raise DataValidationError(
                [
                    "Mapped 'discount' mixes values with and without '%' markers; "
                    "normalize the entire column to one scale."
                ]
            )
    finite = parsed.dropna()
    above_one = finite.gt(1).any()
    fractional = (finite.gt(0) & finite.lt(1)).any()
    if above_one and fractional and not percent_header and not percent_markers.any():
        raise DataValidationError(
            [
                "Mapped 'discount' mixes decimal fractions with whole-number "
                "percentages; normalize the entire column to one scale."
            ]
        )
    percentage_scale = percent_header or percent_markers.any() or above_one
    return (parsed / 100 if percentage_scale else parsed), percentage_scale


def _parse_numbers(series: pd.Series) -> pd.Series:
    normalized = series.map(_normalize_numeric_text)
    return pd.to_numeric(normalized, errors="coerce")


def _parse_dates(series: pd.Series) -> pd.Series:
    values = series.copy()
    text = values.astype("string").str.strip()
    result = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    numeric = pd.to_numeric(text, errors="coerce")
    compact = numeric.between(19_000_101, 21_001_231) & text.str.fullmatch(
        r"\d{8}", na=False
    )
    if compact.any():
        result.loc[compact] = pd.to_datetime(
            text.loc[compact], format="%Y%m%d", errors="coerce"
        )
    excel_serial = (
        numeric.between(20_000, 80_000)
        & text.str.fullmatch(r"\d{5}(?:\.0+)?", na=False)
        & ~compact
    )
    if excel_serial.any():
        result.loc[excel_serial] = pd.to_datetime(
            numeric.loc[excel_serial], unit="D", origin="1899-12-30", errors="coerce"
        )
    remaining = result.isna() & text.ne("")
    if remaining.any():
        parsed = pd.to_datetime(
            text.loc[remaining], errors="coerce", format="mixed", utc=True
        )
        result.loc[remaining] = parsed.dt.tz_localize(None)
    return result


def _parse_import_dates(series: pd.Series) -> pd.Series:
    text = series.astype("string").str.strip()
    timezone_aware = text.str.contains(
        r"(?:[T ]\d{1,2}:\d{2}(?::\d{2}(?:\.\d+)?)?)(?:Z|[+-]\d{2}(?::?\d{2})?)$",
        case=False,
        regex=True,
        na=False,
    )
    if timezone_aware.any():
        raise DataValidationError(
            [
                "Mapped order dates contain "
                f"{int(timezone_aware.sum())} timezone-aware values; provide local "
                "calendar dates without offsets."
            ]
        )
    ambiguous = text.str.extract(
        r"^(?P<first>\d{1,2})(?P<separator>[./-])"
        r"(?P<second>\d{1,2})(?P=separator)(?P<year>\d{2}|\d{4})"
        r"(?:[ T].*)?$"
    )
    first = pd.to_numeric(ambiguous["first"], errors="coerce")
    second = pd.to_numeric(ambiguous["second"], errors="coerce")
    ambiguous_mask = first.between(1, 12) & second.between(1, 12) & first.ne(second)
    if ambiguous_mask.any():
        raise DataValidationError(
            [
                "Mapped order dates contain "
                f"{int(ambiguous_mask.sum())} ambiguous month/day values; use "
                "unambiguous ISO YYYY-MM-DD dates."
            ]
        )
    iso_calendar = text.str.fullmatch(
        r"\d{4}(?P<separator>[./-])\d{1,2}(?P=separator)\d{1,2}"
        r"(?:[ T]\d{1,2}:\d{2}(?::\d{2}(?:\.\d+)?)?)?",
        na=False,
    )
    compact_calendar = text.str.fullmatch(r"\d{8}", na=False)
    excel_serial = text.str.fullmatch(r"\d{5}(?:\.0+)?", na=False)
    unsupported = text.ne("") & ~(iso_calendar | compact_calendar | excel_serial)
    if unsupported.any():
        raise DataValidationError(
            [
                "Mapped order dates contain "
                f"{int(unsupported.sum())} unsupported values; use explicit ISO "
                "YYYY-MM-DD dates, compact YYYYMMDD dates, or five-digit Excel "
                "serial dates."
            ]
        )
    return _parse_dates(series)


def _sample_value(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (date, datetime, pd.Timestamp)):
        text = value.isoformat()
    else:
        text = str(value)
    return text[:PREVIEW_VALUE_CHARACTERS]


def _column_profile(series: pd.Series) -> dict[str, Any]:
    strings = series.astype("string").str.strip()
    non_empty = strings.notna() & strings.ne("")
    populated = series[non_empty]
    count = int(non_empty.sum())
    if count == 0:
        inferred_type = "empty"
    else:
        numeric_ratio = float(_parse_numbers(populated).notna().mean())
        try:
            strict_dates = _parse_import_dates(populated)
            date_ratio = float(strict_dates.notna().mean())
        except DataValidationError:
            date_ratio = 0.0
        if date_ratio >= 0.9 and numeric_ratio < 0.9:
            inferred_type = "date"
        elif numeric_ratio >= 0.9:
            inferred_type = "number"
        elif numeric_ratio >= 0.4 or date_ratio >= 0.4:
            inferred_type = "mixed"
        else:
            inferred_type = "text"
    samples = [_sample_value(value) for value in populated.head(3).tolist()]
    return {
        "inferred_type": inferred_type,
        "non_empty_count": count,
        "unique_count": int(populated.astype("string").nunique(dropna=True)),
        "samples": samples,
    }


def _mapping_suggestions(
    frame: pd.DataFrame, profiles: dict[str, dict[str, Any]]
) -> dict[str, dict[str, Any] | None]:
    normalized_headers = {
        _header_match_key(column): str(column) for column in frame.columns
    }
    suggestions: dict[str, dict[str, Any] | None] = {}
    claimed: set[str] = set()
    for field in MAPPING_FIELDS:
        match = next(
            (
                normalized_headers[_header_match_key(alias)]
                for alias in FIELD_ALIASES[field]
                if _header_match_key(alias) in normalized_headers
            ),
            None,
        )
        if match is not None:
            confidence = 1.0 if _canonical_header(match) == field else 0.95
            suggestions[field] = {
                "column": match,
                "confidence": confidence,
                "reason": "Header matched a known sales-field alias.",
            }
            claimed.add(match)
        else:
            suggestions[field] = None

    if suggestions["order_date"] is None:
        candidates = [
            column
            for column, profile in profiles.items()
            if profile["inferred_type"] == "date" and column not in claimed
        ]
        if len(candidates) == 1:
            column = candidates[0]
            suggestions["order_date"] = {
                "column": column,
                "confidence": 0.72,
                "reason": "Column values consistently parse as dates.",
            }
            claimed.add(column)

    if suggestions["order_id"] is None:
        candidates = []
        for column, profile in profiles.items():
            if column in claimed or profile["non_empty_count"] == 0:
                continue
            unique_ratio = profile["unique_count"] / profile["non_empty_count"]
            if unique_ratio == 1 and profile["inferred_type"] == "text":
                candidates.append(column)
        if len(candidates) == 1:
            suggestions["order_id"] = {
                "column": candidates[0],
                "confidence": 0.55,
                "reason": "Text values are populated and unique; confirm before import.",
            }
    if suggestions["currency"] is None:
        currency_candidates = [
            column
            for column in _currency_candidate_columns(frame)
            if column not in claimed
        ]
        if len(currency_candidates) == 1:
            suggestions["currency"] = {
                "column": currency_candidates[0],
                "confidence": 0.85,
                "reason": "Header or values consistently identify a currency code.",
            }
    return suggestions


def build_preview(dataset: TabularDataset) -> dict[str, Any]:
    profiles = {
        str(column): _column_profile(dataset.frame[column])
        for column in dataset.frame.columns
    }
    columns = [
        {"name": column, **profiles[column]} for column in map(str, dataset.frame.columns)
    ]
    sample_rows = [
        {str(column): _sample_value(value) for column, value in row.items()}
        for row in dataset.frame.head(PREVIEW_ROWS).to_dict(orient="records")
    ]
    warnings: list[str] = []
    if dataset.file_format == "xlsx" and len(dataset.sheets) > 1:
        warnings.append(
            "Only the selected worksheet will be imported; review the worksheet name."
        )
    if any(profile["inferred_type"] == "mixed" for profile in profiles.values()):
        warnings.append(
            "Some columns contain mixed value types and may need a different mapping."
        )
    suggestions = _mapping_suggestions(dataset.frame, profiles)
    suggested_date = suggestions.get("order_date")
    if suggested_date is not None:
        source = suggested_date["column"]
        try:
            _parse_import_dates(dataset.frame[source])
        except DataValidationError as exc:
            warnings.append(
                f"Suggested date column '{source}' needs normalization before import: "
                f"{exc.issues[0]}"
            )
    return {
        "filename": dataset.filename,
        "file_format": dataset.file_format,
        "file_sha256": dataset.file_sha256,
        "sheets": dataset.sheets,
        "selected_sheet": dataset.selected_sheet,
        "row_count": len(dataset.frame),
        "columns": columns,
        "sample_rows": sample_rows,
        "field_definitions": list(FIELD_DEFINITIONS),
        "suggestions": suggestions,
        "warnings": warnings,
    }


def parse_mapping(value: str) -> dict[str, str]:
    try:
        payload = json.loads(value)
    except (json.JSONDecodeError, RecursionError) as exc:
        raise ValueError("mapping must be a valid JSON object.") from exc
    if not isinstance(payload, dict):
        raise ValueError("mapping must be a JSON object.")
    unknown = sorted(set(payload).difference(MAPPING_FIELDS))
    if unknown:
        raise ValueError(f"Unknown mapping fields: {', '.join(unknown)}.")
    mapping: dict[str, str] = {}
    for field, source in payload.items():
        if source is None or source == "":
            continue
        if not isinstance(source, str) or not source.strip():
            raise ValueError(f"Mapping for '{field}' must be a source column name.")
        mapping[field] = source.strip()
    duplicate_sources = sorted(
        source for source, count in Counter(mapping.values()).items() if count > 1
    )
    if duplicate_sources:
        raise DataValidationError(
            [
                "A source column cannot be mapped more than once; repeated "
                f"assignments: {', '.join(duplicate_sources)}."
            ]
        )
    if "order_date" not in mapping:
        raise DataValidationError(["Mapping must include 'order_date'."])
    has_revenue = "revenue" in mapping
    has_components = "quantity" in mapping and "unit_price" in mapping
    if not has_revenue and not has_components:
        raise DataValidationError(
            [
                "Mapping must include either 'revenue' or both 'quantity' and "
                "'unit_price'."
            ]
        )
    if has_revenue and ("unit_price" in mapping or "discount" in mapping):
        raise DataValidationError(
            [
                "Direct revenue mapping cannot also map 'unit_price' or 'discount'; "
                "revenue is already the net amount."
            ]
        )
    return mapping


def _identity_values(
    frame: pd.DataFrame, mapping: dict[str, str], field: str, default: str
) -> tuple[pd.Series, int]:
    source = mapping.get(field)
    if source is None:
        return pd.Series(default, index=frame.index, dtype="string"), len(frame)
    values = frame[source].astype("string").str.strip()
    blank = values.isna() | values.eq("")
    values = values.mask(blank, default)
    return values, int(blank.sum())


def _unique_order_ids(
    frame: pd.DataFrame, source: str | None, file_sha256: str
) -> tuple[pd.Series, int, int]:
    if source is None:
        generated = [
            f"AUTO-{file_sha256[:12]}-{row_number:09d}"
            for row_number in range(1, len(frame) + 1)
        ]
        return pd.Series(generated, index=frame.index, dtype="string"), len(frame), 0
    values = frame[source].astype("string").str.strip()
    blank = values.isna() | values.eq("")
    duplicate = values.notna() & values.ne("") & values.duplicated(keep=False)
    result = values.copy()
    for row_number, index in enumerate(frame.index, start=1):
        if bool(blank.loc[index]):
            result.loc[index] = f"AUTO-{file_sha256[:12]}-{row_number:09d}"
        elif bool(duplicate.loc[index]):
            suffix = f"-ROW-{row_number:09d}"
            maximum = IDENTITY_MAX_LENGTHS["order_id"] - len(suffix)
            result.loc[index] = f"{values.loc[index][:maximum]}{suffix}"
    return result, int(blank.sum()), int(duplicate.sum())


def _entity_semantics(source: str | None) -> tuple[str, str]:
    if source is None:
        return "Unspecified entities", "entities"
    key = _header_match_key(source)
    if any(
        token in key
        for token in ("customer", "client", "buyer", "member", "客户", "客戶")
    ):
        return "Customers", "customers"
    if any(
        token in key
        for token in ("store", "shop", "branch", "门店", "門店")
    ):
        return "Stores", "stores"
    if any(token in key for token in ("account", "账户", "帳戶")):
        return "Accounts", "accounts"
    return "Entities", "entities"


def _record_semantics(source: str | None, *, generated_or_repeated: bool) -> dict[str, Any]:
    if generated_or_repeated:
        return {
            "aggregate_record_proxy": True,
            "record_count_label": "Sales records",
            "average_value_label": "Average sales record value",
            "average_frequency_label": "Average sales records",
            "warning": (
                "Generated or disambiguated identifiers represent imported sales "
                "records, not confirmed source orders."
            ),
        }

    key = _header_match_key(source or "")
    if any(token in key for token in ("order", "invoice", "订单", "訂單", "发票", "發票")):
        labels = ("Orders", "Average order value", "Average orders")
    elif any(token in key for token in ("transaction", "交易")):
        labels = ("Transactions", "Average transaction value", "Average transactions")
    elif any(token in key for token in ("receipt", "收据", "收據")):
        labels = ("Receipts", "Average receipt value", "Average receipts")
    else:
        labels = (
            "Sales records",
            "Average sales record value",
            "Average sales records",
        )
    return {
        "aggregate_record_proxy": False,
        "record_count_label": labels[0],
        "average_value_label": labels[1],
        "average_frequency_label": labels[2],
        "warning": None,
    }


def map_tabular_sales(
    dataset: TabularDataset,
    mapping: dict[str, str],
    *,
    source_currency: str = "USD",
) -> MappedSalesDataset:
    if source_currency not in {"USD", "GBP"}:
        raise ValueError("source_currency must be USD or GBP.")
    missing_sources = sorted(set(mapping.values()).difference(map(str, dataset.frame.columns)))
    if missing_sources:
        raise DataValidationError(
            [f"Mapped source columns do not exist: {', '.join(missing_sources)}."]
        )
    frame = dataset.frame
    currency_candidates = _currency_candidate_columns(frame)
    mapped_currency = mapping.get("currency")
    mapped_non_currency_sources = {
        source for field, source in mapping.items() if field != "currency"
    }
    currency_alias_keys = {
        _header_match_key(alias) for alias in FIELD_ALIASES["currency"]
    }
    unconsumed_currency_columns = [
        column
        for column in currency_candidates
        if column != mapped_currency
        and (
            _header_match_key(column) in currency_alias_keys
            or _strong_currency_header(column)
            or column not in mapped_non_currency_sources
        )
    ]
    if unconsumed_currency_columns:
        raise DataValidationError(
            [
                "Currency-code columns were detected but not mapped: "
                f"{', '.join(unconsumed_currency_columns)}. Map the single source "
                "currency column and remove redundant currency columns before import."
            ]
        )
    if "currency" in mapping:
        _validate_currency_column(
            frame[mapping["currency"]], source_currency=source_currency
        )
    canonical = pd.DataFrame(index=frame.index)
    canonical["order_date"] = _parse_import_dates(frame[mapping["order_date"]])

    order_ids, blank_order_ids, repeated_order_ids = _unique_order_ids(
        frame, mapping.get("order_id"), dataset.file_sha256
    )
    canonical["order_id"] = order_ids
    customer_ids, blank_customers = _identity_values(
        frame, mapping, "customer_id", "UNSPECIFIED-ENTITY"
    )
    canonical["customer_id"] = customer_ids
    canonical["region"], blank_regions = _identity_values(
        frame, mapping, "region", "All regions"
    )
    canonical["category"], blank_categories = _identity_values(
        frame, mapping, "category", "Uncategorized"
    )
    canonical["product"], blank_products = _identity_values(
        frame, mapping, "product", "Unspecified product"
    )

    warnings: list[str] = []
    generated_fields = [
        field for field in MAPPING_FIELDS if field != "currency" and field not in mapping
    ]
    if (blank_order_ids or repeated_order_ids) and "order_id" not in generated_fields:
        generated_fields.append("order_id")
    metric_mode = "direct_revenue" if "revenue" in mapping else "components"
    if metric_mode == "direct_revenue":
        _reject_percent_markers(frame[mapping["revenue"]], field="revenue")
        _validate_monetary_currency(
            frame[mapping["revenue"]],
            field="revenue",
            source_currency=source_currency,
        )
        canonical["revenue"] = _parse_numbers(frame[mapping["revenue"]])
        if "quantity" in mapping:
            _reject_currency_markers(frame[mapping["quantity"]], field="quantity")
            _reject_percent_markers(frame[mapping["quantity"]], field="quantity")
            canonical["quantity"] = _parse_numbers(frame[mapping["quantity"]])
        else:
            canonical["quantity"] = 1
        with np.errstate(divide="ignore", invalid="ignore", over="ignore"):
            canonical["unit_price"] = canonical["revenue"] / canonical["quantity"]
        canonical["discount"] = 0.0
        validated = SalesFrameValidator().validate(
            canonical, preserve_revenue=True
        )
        warnings.append(
            "Direct mapped revenue is treated as the validated net amount and is "
            "not recalculated from price components."
        )
    else:
        _reject_currency_markers(frame[mapping["quantity"]], field="quantity")
        _reject_percent_markers(frame[mapping["quantity"]], field="quantity")
        _reject_percent_markers(frame[mapping["unit_price"]], field="unit_price")
        _validate_monetary_currency(
            frame[mapping["unit_price"]],
            field="unit_price",
            source_currency=source_currency,
        )
        canonical["quantity"] = _parse_numbers(frame[mapping["quantity"]])
        canonical["unit_price"] = _parse_numbers(frame[mapping["unit_price"]])
        if "discount" in mapping:
            _reject_currency_markers(frame[mapping["discount"]], field="discount")
            parsed_discount, percentage_scale = _parse_discount(
                frame[mapping["discount"]], mapping["discount"]
            )
            if percentage_scale:
                warnings.append(
                    "The entire discount column was interpreted as percentages."
                )
            canonical["discount"] = parsed_discount
        else:
            canonical["discount"] = 0.0
        validated = SalesFrameValidator().validate(
            canonical, preserve_component_precision=True
        )

    defaults_used = {
        "order_id": blank_order_ids,
        "customer_id": blank_customers,
        "region": blank_regions,
        "category": blank_categories,
        "product": blank_products,
    }
    for field, count in defaults_used.items():
        if count and field not in generated_fields:
            generated_fields.append(field)
        if count and field in mapping:
            warnings.append(
                f"{count:,} blank '{field}' values used the documented default."
            )
    if repeated_order_ids:
        warnings.append(
            f"{repeated_order_ids:,} rows had repeated order IDs and received "
            "deterministic row suffixes."
        )
    if "order_id" not in mapping:
        warnings.append(
            "Order IDs were generated per row; record counts represent sales records, "
            "not confirmed source orders."
        )
    if "customer_id" not in mapping:
        warnings.append(
            "Customer or entity IDs were not mapped; all rows use one Unspecified "
            "entity and segmentation will be unavailable."
        )

    generated_or_repeated = (
        "order_id" not in mapping or blank_order_ids > 0 or repeated_order_ids > 0
    )
    record_semantics = _record_semantics(
        mapping.get("order_id"), generated_or_repeated=generated_or_repeated
    )
    has_entity_mapping = "customer_id" in mapping and blank_customers < len(frame)
    entity_label, entity_noun = _entity_semantics(
        mapping.get("customer_id") if has_entity_mapping else None
    )
    entity_singular = {
        "customers": "customer",
        "stores": "store",
        "accounts": "account",
        "entities": "entity",
    }.get(entity_noun, "entity")
    if has_entity_mapping and blank_customers:
        entity_warning = (
            f"{blank_customers:,} rows without a {entity_singular} ID were grouped "
            "under UNSPECIFIED-ENTITY and are excluded from entity segmentation."
        )
    elif not has_entity_mapping:
        entity_warning = (
            "No usable customer or entity field was mapped; entity segmentation is "
            "not meaningful for this dataset."
        )
    else:
        entity_warning = None
    units_available = "quantity" in mapping
    unit_warning = (
        None
        if units_available
        else (
            "Source quantity was not mapped. A technical quantity of 1 is stored per "
            "row for schema compatibility and is excluded from unit metrics and "
            "anomaly features."
        )
    )
    anomaly_features = [
        field
        for field in ("revenue", "quantity", "unit_price", "discount")
        if field == "revenue" or field in mapping
    ]
    semantics = {
        **record_semantics,
        "entity_count_label": (
            entity_label if has_entity_mapping else "Unspecified entities"
        ),
        "entity_warning": entity_warning,
        "units_available": units_available,
        "units_label": "Units sold",
        "unit_warning": unit_warning,
        "anomaly_features": anomaly_features,
    }
    return MappedSalesDataset(
        frame=validated,
        mapping=mapping,
        metric_mode=metric_mode,
        generated_fields=sorted(set(generated_fields)),
        warnings=warnings,
        record_semantics=semantics,
    )
