from __future__ import annotations

import hashlib
import json
import zipfile
from dataclasses import replace
from datetime import datetime, timedelta
from io import BytesIO

import pandas as pd
import pytest
from openpyxl import Workbook
from sqlalchemy import select

from backend.app.api import data as data_api
from backend.app.database import SessionLocal
from backend.app.models import SalesRecord


def _delimited_bytes(*, delimiter: str = ",") -> bytes:
    rows = [
        [
            "Transaction Date",
            "Net Sales",
            "Invoice Number",
            "Client",
            "Territory",
            "Department",
            "Item Name",
        ],
        ["2026-01-01", "10.25", "INV-1", "C-1", "North", "Hardware", "Widget"],
        ["2026-02-01", "20.75", "INV-2", "C-2", "South", "Software", "App"],
    ]
    return ("\n".join(delimiter.join(row) for row in rows) + "\n").encode()


def _xlsx_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
    return output.getvalue()


def _replace_xlsx_xml(
    content: bytes, member_name: str, replacements: dict[bytes, bytes]
) -> bytes:
    output = BytesIO()
    applied = {source: False for source in replacements}
    with zipfile.ZipFile(BytesIO(content)) as source_archive:
        with zipfile.ZipFile(output, "w") as target_archive:
            for member in source_archive.infolist():
                payload = source_archive.read(member.filename)
                if member.filename == member_name:
                    for source, replacement in replacements.items():
                        if source in payload:
                            payload = payload.replace(source, replacement)
                            applied[source] = True
                target_archive.writestr(member, payload)
    assert all(applied.values())
    return output.getvalue()


def _post_preview(
    client,
    content: bytes,
    *,
    filename: str = "sales.csv",
    content_type: str = "text/csv",
    sheet_name: str | None = None,
):
    data = {} if sheet_name is None else {"sheet_name": sheet_name}
    return client.post(
        "/api/data/preview",
        data=data,
        files={"file": (filename, content, content_type)},
    )


def _post_import(
    client,
    content: bytes,
    mapping: dict[str, str],
    *,
    filename: str = "sales.csv",
    content_type: str = "text/csv",
    dataset_name: str = "Flexible sales",
    sheet_name: str | None = None,
    expected_sha256: str | None = None,
    source_currency: str = "USD",
    source_profile: str = "order_level",
):
    data = {
        "mapping": json.dumps(mapping),
        "dataset_name": dataset_name,
        "expected_sha256": expected_sha256 or hashlib.sha256(content).hexdigest(),
        "source_currency": source_currency,
        "source_profile": source_profile,
    }
    if sheet_name is not None:
        data["sheet_name"] = sheet_name
    return client.post(
        "/api/data/import",
        data=data,
        files={"file": (filename, content, content_type)},
    )


def test_profile_is_absent_before_any_dataset_is_loaded(client) -> None:
    response = client.get("/api/data/profile")

    assert response.status_code == 404


def test_preview_infers_aliases_and_never_mutates_active_data(client) -> None:
    assert client.post("/api/data/demo").status_code == 200
    before_kpis = client.get("/api/analytics/kpis").json()
    before_profile = client.get("/api/data/profile").json()

    response = _post_preview(client, _delimited_bytes())

    assert response.status_code == 200
    payload = response.json()
    assert payload["filename"] == "sales.csv"
    assert payload["file_format"] == "csv"
    assert len(payload["file_sha256"]) == 64
    assert payload["row_count"] == 2
    assert [column["name"] for column in payload["columns"]] == [
        "Transaction Date",
        "Net Sales",
        "Invoice Number",
        "Client",
        "Territory",
        "Department",
        "Item Name",
    ]
    suggestion_text = json.dumps(payload["suggestions"]).lower()
    assert "order_date" in suggestion_text
    assert "transaction date" in suggestion_text
    assert "revenue" in suggestion_text
    assert "net sales" in suggestion_text
    assert payload["sample_rows"]
    assert payload["field_definitions"]
    assert client.get("/api/analytics/kpis").json() == before_kpis
    assert client.get("/api/data/profile").json() == before_profile


def test_direct_revenue_import_preserves_totals_and_discloses_defaults(client) -> None:
    content = (
        b"When,Amount,Invoice\n"
        b"2026-01-01,10.25,REPEATED\n"
        b"2026-02-01,20.75,REPEATED\n"
    )
    mapping = {
        "order_date": "When",
        "revenue": "Amount",
        "order_id": "Invoice",
    }

    response = _post_import(client, content, mapping, dataset_name="Direct totals")

    assert response.status_code == 200
    payload = response.json()
    assert payload["rows_loaded"] == 2
    assert payload["revenue_total"] == 31.0
    assert payload["mapping"] == mapping
    assert "order_id" in payload["generated_fields"]
    assert "customer_id" in payload["generated_fields"]
    dataset_profile = payload["dataset_profile"]
    assert dataset_profile["currency"] == "USD"
    assert dataset_profile["units_available"] is False
    assert dataset_profile["units_label"]
    assert dataset_profile["unit_warning"]
    assert "revenue" in dataset_profile["anomaly_features"]
    assert not {
        "quantity",
        "unit_price",
        "discount",
    }.intersection(dataset_profile["anomaly_features"])
    with SessionLocal() as session:
        records = list(session.scalars(select(SalesRecord).order_by(SalesRecord.id)))
    assert [record.revenue for record in records] == [10.25, 20.75]
    assert [record.quantity for record in records] == [1, 1]
    assert [record.discount for record in records] == [0.0, 0.0]
    assert len({record.order_id for record in records}) == 2
    assert {record.customer_id for record in records} == {"UNSPECIFIED-ENTITY"}

    profile = client.get("/api/data/profile")
    assert profile.status_code == 200
    profile_text = json.dumps(profile.json()).lower()
    assert "direct totals" in profile_text
    assert "sales records" in profile_text
    assert "unspecified" in profile_text

    dashboard = client.get("/api/dashboard").json()
    assert dashboard["kpis"]["record_semantics"]["record_count_label"] == "Sales records"
    assert dashboard["kpis"]["units_sold"] is None
    assert dashboard["segments"] is None
    assert "segments" in dashboard["model_errors"]


def test_quantity_price_mapping_derives_revenue(client) -> None:
    content = (
        b"Sold On,Units,List Price,Discount Percent,SKU\n"
        b"2026-01-01,2,10,10,WIDGET\n"
        b"2026-02-01,3,20,0,APP\n"
    )
    mapping = {
        "order_date": "Sold On",
        "quantity": "Units",
        "unit_price": "List Price",
        "discount": "Discount Percent",
        "product": "SKU",
    }

    response = _post_import(client, content, mapping, dataset_name="Component facts")

    assert response.status_code == 200
    assert response.json()["revenue_total"] == 78.0
    assert response.json()["dataset_profile"]["units_available"] is True
    assert response.json()["dataset_profile"]["unit_warning"] is None
    assert "quantity" in response.json()["dataset_profile"]["anomaly_features"]
    assert any(
        "entire discount column" in warning.lower()
        for warning in response.json()["warnings"]
    )
    with SessionLocal() as session:
        revenues = list(
            session.scalars(select(SalesRecord.revenue).order_by(SalesRecord.id))
        )
    assert revenues == [18.0, 60.0]


@pytest.mark.parametrize("malformed_value", ['"1,2,3"', "1$2", "USD 1 2"])
def test_import_rejects_malformed_mapped_numeric_strings(
    client, malformed_value
) -> None:
    content = f"When,Amount\n2026-01-01,{malformed_value}\n".encode()

    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
    )

    assert response.status_code == 422
    assert "non-numeric" in " ".join(response.json()["detail"]).lower()


def test_component_import_preserves_sub_cent_price_before_revenue_rounding(
    client,
) -> None:
    content = b"When,Units,Price\n2026-01-01,1000,0.0049\n"

    response = _post_import(
        client,
        content,
        {
            "order_date": "When",
            "quantity": "Units",
            "unit_price": "Price",
        },
    )

    assert response.status_code == 200
    assert response.json()["revenue_total"] == pytest.approx(4.90)
    with SessionLocal() as session:
        record = session.scalar(select(SalesRecord))
    assert record is not None
    assert record.quantity == 1000
    assert record.unit_price == pytest.approx(0.0049)
    assert record.revenue == pytest.approx(4.90)


@pytest.mark.parametrize(
    ("source_currency", "amount"),
    [("USD", "$12.50"), ("GBP", "£12.50")],
)
def test_source_currency_is_persisted_in_import_and_profile(
    client, source_currency, amount
) -> None:
    content = f"When,Amount\n2026-01-01,{amount}\n".encode()

    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
        source_currency=source_currency,
    )

    assert response.status_code == 200
    assert response.json()["dataset_profile"]["currency"] == source_currency
    profile = client.get("/api/data/profile")
    assert profile.status_code == 200
    assert profile.json()["currency"] == source_currency


def test_dataset_version_changes_when_same_file_uses_a_different_mapping(client) -> None:
    content = (
        b"When,Amount A,Amount B\n"
        b"2026-01-01,10,100\n"
        b"2026-02-01,20,200\n"
    )
    first = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount A"},
    )
    first_version = client.get("/api/dashboard").json()["dataset_version"]

    second = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount B"},
    )
    second_version = client.get("/api/dashboard").json()["dataset_version"]

    assert first.status_code == 200
    assert second.status_code == 200
    assert first_version["content_sha256"] == second_version["content_sha256"]
    assert first_version["profile_sha256"] != second_version["profile_sha256"]
    assert first.json()["revenue_total"] == pytest.approx(30.0)
    assert second.json()["revenue_total"] == pytest.approx(300.0)


@pytest.mark.parametrize(
    (
        "source_profile",
        "source_currency",
        "expected_name",
        "expected_records",
        "expected_entities",
    ),
    [
        (
            "m5",
            "USD",
            "Walmart M5",
            "Store-category-day records",
            "Stores",
        ),
        (
            "uci",
            "GBP",
            "UCI Online Retail II",
            "Customer-country-day records",
            "Customers",
        ),
        (
            "iowa",
            "USD",
            "Iowa Liquor Sales 2024",
            "Store-county-category-month records",
            "Stores",
        ),
    ],
)
def test_prepared_source_profile_preserves_aggregate_business_meaning(
    client,
    monkeypatch,
    source_profile,
    source_currency,
    expected_name,
    expected_records,
    expected_entities,
) -> None:
    if source_profile == "m5":
        rows = [
            "M5-d_1-CA_1-FOODS,2011-01-29,CA_1,CA,FOODS,CA_1 FOODS,2,5,0",
            "M5-d_2-TX_1-HOBBIES,2011-01-30,TX_1,TX,HOBBIES,TX_1 HOBBIES,3,10,0",
        ]
    elif source_profile == "uci":
        rows = [
            "UCI-00000001,2009-12-01,UCI-12345,UK,Online Retail,Daily online retail basket,2,5,0",
            "UCI-00000002,2009-12-02,UCI-99999,France,Online Retail,"
            "Daily online retail basket,3,10,0",
        ]
    else:
        rows = [
            "IA2024-00000001,2024-01-01,IA-STORE-1,Polk,Whiskey,Monthly spirits basket,2,5,0",
            "IA2024-00000002,2024-02-01,IA-STORE-2,Linn,Vodka,Monthly spirits basket,3,10,0",
        ]
    content = (
        "order_id,order_date,customer_id,region,category,product,quantity,unit_price,discount\n"
        + "\n".join(rows)
        + "\n"
    ).encode()
    mapping = {field: field for field in data_api.REQUIRED_COLUMNS}
    observed_dates = [row.split(",")[1] for row in rows]
    monkeypatch.setitem(
        data_api.KNOWN_SOURCE_EXPECTED_SUMMARIES,
        source_profile,
        (
            2,
            datetime.fromisoformat(min(observed_dates)).date(),
            datetime.fromisoformat(max(observed_dates)).date(),
        ),
    )

    response = _post_import(
        client,
        content,
        mapping,
        source_profile=source_profile,
        source_currency=source_currency,
    )

    assert response.status_code == 200
    profile = response.json()["dataset_profile"]
    assert profile["dataset_name"] == expected_name
    assert profile["currency"] == source_currency
    assert profile["record_count_label"] == expected_records
    assert profile["entity_count_label"] == expected_entities
    assert profile["aggregate_record_proxy"] is True


def test_prepared_profile_rejects_rows_without_source_entity_ids(client) -> None:
    content = (
        b"order_id,order_date,customer_id,region,category,product,quantity,unit_price,discount\n"
        b"M5-d_1-CA_1-FOODS,2011-01-29,,CA,FOODS,CA_1 FOODS,2,5,0\n"
        b"M5-d_2-TX_1-HOBBIES,2011-01-30,TX_1,TX,HOBBIES,TX_1 HOBBIES,3,10,0\n"
    )

    response = _post_import(
        client,
        content,
        {field: field for field in data_api.REQUIRED_COLUMNS},
        source_profile="m5",
    )

    assert response.status_code == 422
    assert "requires a source entity ID for every row" in " ".join(
        response.json()["detail"]
    )


@pytest.mark.parametrize(
    ("source_profile", "source_currency", "row"),
    [
        (
            "uci",
            "GBP",
            "UCI-00000001,2009-12-01,UCI-12345,,Online Retail,"
            "Daily online retail basket,2,5,0",
        ),
        (
            "iowa",
            "USD",
            "IA2024-00000001,2024-01-01,IA-STORE-1,,,"
            "Monthly spirits basket,2,5,0",
        ),
    ],
)
def test_prepared_profile_rejects_generic_dimension_defaults(
    client, monkeypatch, source_profile, source_currency, row
) -> None:
    content = (
        "order_id,order_date,customer_id,region,category,product,"
        f"quantity,unit_price,discount\n{row}\n"
    ).encode()
    observed_date = datetime.fromisoformat(row.split(",")[1]).date()
    monkeypatch.setitem(
        data_api.KNOWN_SOURCE_EXPECTED_SUMMARIES,
        source_profile,
        (1, observed_date, observed_date),
    )

    response = _post_import(
        client,
        content,
        {field: field for field in data_api.REQUIRED_COLUMNS},
        source_profile=source_profile,
        source_currency=source_currency,
    )

    assert response.status_code == 422
    assert "outside its documented" in " ".join(response.json()["detail"])


def test_prepared_profile_requires_the_complete_expected_contract(client) -> None:
    content = (
        b"order_id,order_date,customer_id,region,category,product,quantity,unit_price,discount\n"
        b"UCI-00000001,2009-12-01,UCI-12345,UK,Online Retail,Daily online retail basket,2,5,0\n"
    )

    response = _post_import(
        client,
        content,
        {field: field for field in data_api.REQUIRED_COLUMNS},
        source_profile="uci",
        source_currency="GBP",
    )

    assert response.status_code == 422
    assert "requires the complete prepared-output contract" in " ".join(
        response.json()["detail"]
    )


def test_import_response_keeps_the_profile_written_by_its_transaction(
    client, monkeypatch
) -> None:
    original_load = data_api.load_sales_frame

    def replace_profile_after_commit(*args, **kwargs):
        result = original_load(*args, **kwargs)
        with SessionLocal() as other_session:
            active = other_session.get(data_api.DatasetProfile, 1)
            active.dataset_name = "Concurrent replacement"
            other_session.commit()
        return result

    monkeypatch.setattr(data_api, "load_sales_frame", replace_profile_after_commit)
    content = b"When,Amount\n2026-01-01,10\n"
    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
        dataset_name="Transaction A",
    )

    assert response.status_code == 200
    assert response.json()["dataset_profile"]["dataset_name"] == "Transaction A"
    assert client.get("/api/data/profile").json()["dataset_name"] == (
        "Concurrent replacement"
    )


@pytest.mark.parametrize(
    ("source_currency", "amounts", "message"),
    [
        ("USD", ["£10"], "selected source currency"),
        ("USD", ["$10", "£20"], "selected source currency"),
        ("GBP", ["€10"], "unsupported currency"),
        ("USD", ["EUR 10"], "unsupported currency"),
    ],
)
def test_import_rejects_mismatched_mixed_or_unsupported_currency_markers(
    client, source_currency, amounts, message
) -> None:
    rows = "".join(
        f"2026-{index:02d}-01,{amount}\n"
        for index, amount in enumerate(amounts, start=1)
    )
    content = f"When,Amount\n{rows}".encode()

    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
        source_currency=source_currency,
    )

    assert response.status_code == 422
    assert message in " ".join(response.json()["detail"]).lower()


@pytest.mark.parametrize(
    ("currency_values", "message"),
    [
        (["USD", "GBP"], "does not match"),
        (["USD", "EUR"], "unsupported"),
        (["USD", ""], "blank"),
    ],
)
def test_mapped_currency_column_rejects_mixed_unsupported_or_blank_values(
    client, currency_values, message
) -> None:
    content = (
        "When,Amount,Currency\n"
        f"2026-01-01,10,{currency_values[0]}\n"
        f"2026-02-01,20,{currency_values[1]}\n"
    ).encode()

    response = _post_import(
        client,
        content,
        {
            "order_date": "When",
            "revenue": "Amount",
            "currency": "Currency",
        },
        source_currency="USD",
    )

    assert response.status_code == 422
    assert message in " ".join(response.json()["detail"]).lower()


def test_mapped_all_usd_currency_column_succeeds_and_is_persisted(client) -> None:
    content = (
        b"When,Amount,Currency\n"
        b"2026-01-01,10,USD\n"
        b"2026-02-01,20,$\n"
    )
    mapping = {
        "order_date": "When",
        "revenue": "Amount",
        "currency": "Currency",
    }

    response = _post_import(
        client,
        content,
        mapping,
        source_currency="USD",
    )

    assert response.status_code == 200
    profile = response.json()["dataset_profile"]
    assert profile["currency"] == "USD"
    assert profile["mapped_fields"]["currency"] == "Currency"
    persisted = client.get("/api/data/profile")
    assert persisted.status_code == 200
    assert persisted.json()["currency"] == "USD"
    assert persisted.json()["mapped_fields"]["currency"] == "Currency"


def test_obvious_currency_column_cannot_be_silently_ignored(client) -> None:
    content = b"When,Amount,Currency\n2026-01-01,10,USD\n"

    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
    )

    assert response.status_code == 422
    assert "currency-code column" in " ".join(response.json()["detail"]).lower()
    assert "not mapped" in " ".join(response.json()["detail"]).lower()


def test_transaction_currency_cannot_be_omitted_for_mixed_currency_rows(client) -> None:
    content = (
        b"When,Amount,transaction_currency\n"
        b"2026-01-01,10,USD\n"
        b"2026-02-01,20,GBP\n"
    )

    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
        source_currency="USD",
    )

    assert response.status_code == 422
    detail = " ".join(response.json()["detail"]).lower()
    assert "transaction_currency" in detail
    assert "not mapped" in detail


def test_product_values_that_are_currency_codes_are_not_treated_as_currency(
    client,
) -> None:
    content = (
        b"When,Amount,Product\n"
        b"2026-01-01,10,USD\n"
        b"2026-02-01,20,USD\n"
    )
    mapping = {
        "order_date": "When",
        "revenue": "Amount",
        "product": "Product",
    }

    preview = _post_preview(client, content)
    response = _post_import(client, content, mapping)

    assert preview.status_code == 200
    assert preview.json()["suggestions"]["product"]["column"] == "Product"
    assert preview.json()["suggestions"]["currency"] is None
    assert response.status_code == 200
    with SessionLocal() as session:
        products = list(
            session.scalars(select(SalesRecord.product).order_by(SalesRecord.id))
        )
    assert products == ["USD", "USD"]


@pytest.mark.parametrize(
    ("header", "currency_values"),
    [
        ("transaction_currency", ["CHF"]),
        ("sourceCurrency", ["USD", "GBP", "N/A"]),
        ("local_currency", ["N/A"]),
    ],
)
def test_explicit_currency_code_headers_cannot_be_omitted(
    client, header, currency_values
) -> None:
    rows = "".join(
        f"2026-01-{index:02d},{index * 10},{value}\n"
        for index, value in enumerate(currency_values, start=1)
    )
    content = f"When,Amount,{header}\n{rows}".encode()

    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
    )

    assert response.status_code == 422
    detail = " ".join(response.json()["detail"])
    assert header in detail
    assert "not mapped" in detail.lower()


def test_mapped_chf_currency_reports_the_actual_unsupported_code(client) -> None:
    content = b"When,Amount,transaction_currency\n2026-01-01,10,CHF\n"

    response = _post_import(
        client,
        content,
        {
            "order_date": "When",
            "revenue": "Amount",
            "currency": "transaction_currency",
        },
        source_currency="USD",
    )

    assert response.status_code == 422
    detail = " ".join(response.json()["detail"])
    assert "CHF" in detail
    assert "unsupported" in detail.lower()
    assert "USD or GBP" in detail


def test_currency_words_in_mapped_business_fields_do_not_create_false_positives(
    client,
) -> None:
    product_content = (
        b"When,Amount,Cryptocurrency\n"
        b"2026-01-01,10,USD\n"
        b"2026-02-01,20,GBP\n"
    )
    product_import = _post_import(
        client,
        product_content,
        {
            "order_date": "When",
            "revenue": "Amount",
            "product": "Cryptocurrency",
        },
    )

    assert product_import.status_code == 200
    with SessionLocal() as session:
        assert set(session.scalars(select(SalesRecord.product))) == {"USD", "GBP"}

    revenue_content = (
        b"When,currency_adjusted_revenue\n"
        b"2026-01-01,$10\n"
        b"2026-02-01,$20\n"
    )
    revenue_import = _post_import(
        client,
        revenue_content,
        {
            "order_date": "When",
            "revenue": "currency_adjusted_revenue",
        },
        source_currency="USD",
    )

    assert revenue_import.status_code == 200
    assert revenue_import.json()["revenue_total"] == pytest.approx(30.0)


def test_import_rejects_unsupported_source_currency(client) -> None:
    content = b"When,Amount\n2026-01-01,10\n"

    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
        source_currency="EUR",
    )

    assert response.status_code == 422
    detail = json.dumps(response.json()["detail"]).lower()
    assert "usd" in detail
    assert "gbp" in detail


@pytest.mark.parametrize(
    ("content", "mapping", "source_currency", "expected_marker"),
    [
        (
            b"When,Revenue (USD)\n2026-01-01,10\n",
            {"order_date": "When", "revenue": "Revenue (USD)"},
            "GBP",
            "USD",
        ),
        (
            b"When,Units,Unit Price GBP\n2026-01-01,2,10\n",
            {
                "order_date": "When",
                "quantity": "Units",
                "unit_price": "Unit Price GBP",
            },
            "USD",
            "GBP",
        ),
    ],
)
def test_import_rejects_currency_mismatch_declared_in_monetary_header(
    client, content, mapping, source_currency, expected_marker
) -> None:
    response = _post_import(
        client,
        content,
        mapping,
        source_currency=source_currency,
    )

    assert response.status_code == 422
    detail = " ".join(response.json()["detail"])
    assert expected_marker in detail
    assert source_currency in detail


def test_strong_currency_column_cannot_be_remapped_as_a_business_dimension(client) -> None:
    content = (
        b"When,Amount,Currency Description\n"
        b"2026-01-01,10,CHF\n"
        b"2026-02-01,20,CHF\n"
    )

    response = _post_import(
        client,
        content,
        {
            "order_date": "When",
            "revenue": "Amount",
            "product": "Currency Description",
        },
    )

    assert response.status_code == 422
    assert "currency-code columns" in " ".join(response.json()["detail"]).lower()


@pytest.mark.parametrize(
    ("content", "mapping"),
    [
        (
            b"When,Amount\n2026-01-01,10%\n",
            {"order_date": "When", "revenue": "Amount"},
        ),
        (
            b"When,Units,Price\n2026-01-01,2%,10\n",
            {
                "order_date": "When",
                "quantity": "Units",
                "unit_price": "Price",
            },
        ),
        (
            b"When,Units,Price\n2026-01-01,2,10%\n",
            {
                "order_date": "When",
                "quantity": "Units",
                "unit_price": "Price",
            },
        ),
    ],
)
def test_import_rejects_percent_markers_outside_discount(client, content, mapping) -> None:
    response = _post_import(client, content, mapping)

    assert response.status_code == 422
    assert "percent" in " ".join(response.json()["detail"]).lower()


@pytest.mark.parametrize(
    "discounts",
    [
        ["0.1", "10"],
        ["10%", "20"],
    ],
)
def test_discount_rejects_mixed_fraction_and_percent_scales(client, discounts) -> None:
    content = (
        "When,Units,Price,Discount\n"
        f"2026-01-01,1,100,{discounts[0]}\n"
        f"2026-02-01,1,100,{discounts[1]}\n"
    ).encode()

    response = _post_import(
        client,
        content,
        {
            "order_date": "When",
            "quantity": "Units",
            "unit_price": "Price",
            "discount": "Discount",
        },
    )

    assert response.status_code == 422
    assert "normalize the entire column" in " ".join(response.json()["detail"]).lower()


def test_discount_percent_header_scales_small_decimal_values_as_percentages(
    client,
) -> None:
    content = (
        b"When,Units,Price,Discount %\n"
        b"2026-01-01,1,100,0.1\n"
        b"2026-02-01,1,100,0.2\n"
    )

    response = _post_import(
        client,
        content,
        {
            "order_date": "When",
            "quantity": "Units",
            "unit_price": "Price",
            "discount": "Discount %",
        },
    )

    assert response.status_code == 200
    assert response.json()["revenue_total"] == pytest.approx(199.70)
    with SessionLocal() as session:
        discounts = list(
            session.scalars(select(SalesRecord.discount).order_by(SalesRecord.id))
        )
    assert discounts == pytest.approx([0.001, 0.002])


def test_discount_rejects_repeated_percent_markers(client) -> None:
    content = b"When,Units,Price,Discount %\n2026-01-01,1,100,10%%\n"

    response = _post_import(
        client,
        content,
        {
            "order_date": "When",
            "quantity": "Units",
            "unit_price": "Price",
            "discount": "Discount %",
        },
    )

    assert response.status_code == 422
    assert "malformed percentage notation" in " ".join(
        response.json()["detail"]
    ).lower()


def test_repeated_max_length_source_ids_receive_bounded_generated_ids(client) -> None:
    repeated_id = "X" * 80
    content = (
        "When,Amount,Invoice\n"
        f"2026-01-01,10,{repeated_id}\n"
        f"2026-02-01,20,{repeated_id}\n"
    ).encode()

    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount", "order_id": "Invoice"},
    )

    assert response.status_code == 200
    with SessionLocal() as session:
        order_ids = list(
            session.scalars(select(SalesRecord.order_id).order_by(SalesRecord.id))
        )
    assert len(set(order_ids)) == 2
    assert all(len(order_id) <= 80 for order_id in order_ids)


@pytest.mark.parametrize(
    ("id_column", "record_label", "average_label"),
    [
        ("Transaction ID", "Transactions", "Average transaction value"),
        ("Receipt ID", "Receipts", "Average receipt value"),
    ],
)
def test_unique_transaction_and_receipt_ids_keep_their_source_semantics(
    client, id_column, record_label, average_label
) -> None:
    content = (
        f"When,Amount,{id_column}\n"
        "2026-01-01,10,ID-1\n"
        "2026-02-01,20,ID-2\n"
    ).encode()

    response = _post_import(
        client,
        content,
        {
            "order_date": "When",
            "revenue": "Amount",
            "order_id": id_column,
        },
    )

    assert response.status_code == 200
    profile = response.json()["dataset_profile"]
    assert profile["record_count_label"] == record_label
    assert profile["record_count_label"] != "Orders"
    assert profile["average_value_label"] == average_label
    assert profile["aggregate_record_proxy"] is False


def test_profile_discloses_defaults_used_for_blank_mapped_values(client) -> None:
    content = b"When,Amount,Customer\n2026-01-01,10,\n"

    response = _post_import(
        client,
        content,
        {
            "order_date": "When",
            "revenue": "Amount",
            "customer_id": "Customer",
        },
    )

    assert response.status_code == 200
    assert "customer_id" in response.json()["generated_fields"]
    assert any("blank 'customer_id'" in warning for warning in response.json()["warnings"])
    with SessionLocal() as session:
        assert session.scalar(select(SalesRecord.customer_id)) == "UNSPECIFIED-ENTITY"


def test_partial_unspecified_entities_are_excluded_from_segmentation(client) -> None:
    content = (
        b"When,Amount,Customer\n"
        b"2026-01-01,10,C-1\n"
        b"2026-02-01,20,C-1\n"
        b"2026-01-15,30,C-2\n"
        b"2026-02-15,30,C-2\n"
        b"2026-02-20,1000,\n"
    )

    imported = _post_import(
        client,
        content,
        {
            "order_date": "When",
            "revenue": "Amount",
            "customer_id": "Customer",
        },
    )

    assert imported.status_code == 200
    profile = imported.json()["dataset_profile"]
    assert "excluded from entity segmentation" in profile["entity_warning"].lower()

    response = client.get("/api/ml/segments")
    assert response.status_code == 200
    payload = response.json()
    assert payload["excluded_unspecified_records"] == 1
    assert {
        customer["customer_id"] for customer in payload["customers"]
    } == {"C-1", "C-2"}
    assert sum(segment["total_revenue"] for segment in payload["segments"]) == 90.0


def test_tsv_import_uses_the_same_mapping_contract(client) -> None:
    content = _delimited_bytes(delimiter="\t")
    mapping = {
        "order_date": "Transaction Date",
        "revenue": "Net Sales",
        "order_id": "Invoice Number",
        "customer_id": "Client",
        "region": "Territory",
        "category": "Department",
        "product": "Item Name",
    }

    preview = _post_preview(
        client,
        content,
        filename="sales.tsv",
        content_type="text/tab-separated-values",
    )
    imported = _post_import(
        client,
        content,
        mapping,
        filename="sales.tsv",
        content_type="text/tab-separated-values",
    )

    assert preview.status_code == 200
    assert preview.json()["file_format"] == "tsv"
    assert imported.status_code == 200
    assert imported.json()["rows_loaded"] == 2


def test_chinese_headers_are_suggested_and_imported_without_ascii_rewrites(client) -> None:
    content = (
        "銷售日期,銷售額,訂單號,門店,地區,品類,商品\n"
        "2026-01-01,100,訂單一,台北店,北區,飲品,茶\n"
        "2026-02-01,200,訂單二,高雄店,南區,食品,餅乾\n"
    ).encode()

    preview = _post_preview(client, content, filename="銷售資料.csv")

    assert preview.status_code == 200
    suggestions = preview.json()["suggestions"]
    expected = {
        "order_date": "銷售日期",
        "revenue": "銷售額",
        "order_id": "訂單號",
        "customer_id": "門店",
        "region": "地區",
        "category": "品類",
        "product": "商品",
    }
    assert {
        field: suggestions[field]["column"] for field in expected
    } == expected

    imported = _post_import(
        client,
        content,
        expected,
        filename="銷售資料.csv",
        dataset_name="中文銷售資料",
    )

    assert imported.status_code == 200
    profile = imported.json()["dataset_profile"]
    assert profile["dataset_name"] == "中文銷售資料"
    assert profile["entity_count_label"] == "Stores"
    with SessionLocal() as session:
        assert set(session.scalars(select(SalesRecord.product))) == {"茶", "餅乾"}


@pytest.mark.parametrize(
    ("entity_column", "expected_label", "expected_word", "other_word"),
    [
        ("Store ID", "Stores", "stores", "accounts"),
        ("Account ID", "Accounts", "accounts", "stores"),
    ],
)
def test_store_and_account_entities_are_not_presented_as_customers(
    client, entity_column, expected_label, expected_word, other_word
) -> None:
    content = (
        f"When,Amount,{entity_column}\n"
        "2026-01-01,10,E-1\n"
        "2026-02-01,20,E-2\n"
    ).encode()
    mapping = {
        "order_date": "When",
        "revenue": "Amount",
        "customer_id": entity_column,
    }
    imported = _post_import(client, content, mapping)

    assert imported.status_code == 200
    profile = imported.json()["dataset_profile"]
    assert profile["entity_count_label"] == expected_label
    assert profile["entity_count_label"] != "Customers"

    insight = client.post(
        "/api/insights/query",
        json={"question": "Top customers by revenue"},
    )
    assert insight.status_code == 200
    payload = insight.json()
    assert payload["tools_used"] == []
    assert "unavailable" in payload["answer"].lower()
    assert expected_word in payload["answer"].lower()

    matching = client.post(
        "/api/insights/query",
        json={"question": f"Top {expected_word} by revenue"},
    )
    assert matching.status_code == 200
    matching_payload = matching.json()
    assert matching_payload["tools_used"] == ["approved_analytics_query"]
    assert matching_payload["query_plan"]["dimension"] == "customer"
    assert expected_label in matching_payload["chart"]["title"]
    assert "unavailable" not in matching_payload["answer"].lower()

    generic_matching = client.post(
        "/api/insights/query",
        json={"question": "Top entities by revenue"},
    )
    assert generic_matching.status_code == 200
    generic_payload = generic_matching.json()
    assert generic_payload["tools_used"] == ["approved_analytics_query"]
    assert expected_label in generic_payload["chart"]["title"]
    assert "customer" not in generic_payload["answer"].lower()


    specialist_question = (
        "Segment shops" if expected_label == "Stores" else "Segment accounts"
    )
    specialist = client.post(
        "/api/insights/query",
        json={"question": specialist_question},
    )
    assert specialist.status_code == 200
    assert specialist.json()["tools_used"] == ["customer_segments"]
    assert expected_word in specialist.json()["answer"].lower()

    mismatched = client.post(
        "/api/insights/query",
        json={"question": f"Top {other_word} by revenue"},
    )
    assert mismatched.status_code == 200
    mismatched_payload = mismatched.json()
    assert mismatched_payload["tools_used"] == []
    assert "unavailable" in mismatched_payload["answer"].lower()
    assert expected_word in mismatched_payload["answer"].lower()

    specialist_mismatch = client.post(
        "/api/insights/query",
        json={"question": f"Segment {other_word}"},
    )
    assert specialist_mismatch.status_code == 200
    assert specialist_mismatch.json()["tools_used"] == []
    assert "unavailable" in specialist_mismatch.json()["answer"].lower()


@pytest.mark.parametrize("entity_word", ["entity", "entities"])
def test_generic_entity_questions_are_gracefully_unavailable_without_entity_mapping(
    client, entity_word
) -> None:
    content = b"When,Amount\n2026-01-01,10\n2026-02-01,20\n"
    imported = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
    )
    assert imported.status_code == 200

    response = client.post(
        "/api/insights/query",
        json={"question": f"Top {entity_word} by revenue"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["tools_used"] == []
    assert "unavailable" in payload["answer"].lower()
    assert "unspecified entities" in payload["answer"].lower()


def test_xlsx_preview_and_import_use_the_selected_sheet(client) -> None:
    content = _xlsx_bytes(
        {
            "Instructions": pd.DataFrame({"Read me": ["Choose the Sales sheet"]}),
            "Sales 2026": pd.DataFrame(
                {
                    "Sale Date": [pd.Timestamp("2026-01-01"), pd.Timestamp("2026-02-01")],
                    "Line Total": [12.5, 25.0],
                    "Product Code": ["A", "B"],
                }
            ),
        }
    )
    mapping = {
        "order_date": "Sale Date",
        "revenue": "Line Total",
        "product": "Product Code",
    }

    preview = _post_preview(
        client,
        content,
        filename="sales.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        sheet_name="Sales 2026",
    )
    imported = _post_import(
        client,
        content,
        mapping,
        filename="sales.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        dataset_name="Workbook sales",
        sheet_name="Sales 2026",
    )

    assert preview.status_code == 200
    assert preview.json()["file_format"] == "xlsx"
    assert preview.json()["sheets"] == ["Instructions", "Sales 2026"]
    assert preview.json()["selected_sheet"] == "Sales 2026"
    assert imported.status_code == 200
    assert imported.json()["revenue_total"] == 37.5
    assert imported.json()["dataset_profile"]["source_sheet"] == "Sales 2026"

    missing_sheet = _post_preview(
        client,
        content,
        filename="sales.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        sheet_name="Missing",
    )
    assert missing_sheet.status_code == 400


def test_xlsx_understated_dimension_does_not_truncate_real_columns(client) -> None:
    content = _xlsx_bytes(
        {
            "Sales": pd.DataFrame(
                {
                    "When": ["2026-01-01", "2026-02-01"],
                    "Units": [2, 3],
                    "Price": [10, 20],
                }
            )
        }
    )
    content = _replace_xlsx_xml(
        content,
        "xl/worksheets/sheet1.xml",
        {b'<dimension ref="A1:C3" />': b'<dimension ref="A1:B3" />'},
    )
    mapping = {
        "order_date": "When",
        "quantity": "Units",
        "unit_price": "Price",
    }

    preview = _post_preview(
        client,
        content,
        filename="understated-dimension.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    imported = _post_import(
        client,
        content,
        mapping,
        filename="understated-dimension.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    assert preview.status_code == 200
    assert [column["name"] for column in preview.json()["columns"]] == [
        "When",
        "Units",
        "Price",
    ]
    assert preview.json()["sample_rows"][0]["Price"] == "10"
    assert imported.status_code == 200
    assert imported.json()["revenue_total"] == pytest.approx(80.0)


def test_xlsx_large_adjacent_numeric_ids_and_blank_remain_distinct(client) -> None:
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sales"
    worksheet.append(["When", "Amount", "Transaction ID"])
    worksheet.append(["2026-01-01", 10, 9_007_199_254_740_994])
    worksheet.append(["2026-02-01", 20, 9_007_199_254_740_996])
    worksheet.append(["2026-03-01", 30, None])
    workbook.save(output)
    workbook.close()
    content = _replace_xlsx_xml(
        output.getvalue(),
        "xl/worksheets/sheet1.xml",
        {
            b">9007199254740994<": b">9007199254740993<",
            b">9007199254740996<": b">9007199254740994<",
        },
    )
    mapping = {
        "order_date": "When",
        "revenue": "Amount",
        "order_id": "Transaction ID",
    }

    preview = _post_preview(
        client,
        content,
        filename="large-ids.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    imported = _post_import(
        client,
        content,
        mapping,
        filename="large-ids.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    assert preview.status_code == 200
    assert [
        row["Transaction ID"] for row in preview.json()["sample_rows"][:2]
    ] == ["9007199254740993", "9007199254740994"]
    assert imported.status_code == 200
    with SessionLocal() as session:
        order_ids = list(
            session.scalars(select(SalesRecord.order_id).order_by(SalesRecord.id))
        )
    assert len(order_ids) == 3
    assert len(set(order_ids)) == 3
    assert {"9007199254740993", "9007199254740994"}.issubset(order_ids)
    assert sum(order_id.startswith("AUTO-") for order_id in order_ids) == 1
    assert not any("-ROW-" in order_id for order_id in order_ids)


@pytest.mark.parametrize(
    "mapping",
    [
        {"revenue": "Amount"},
        {"order_date": "When"},
        {
            "order_date": "When",
            "revenue": "Amount",
            "quantity": "Units",
            "unit_price": "Price",
        },
        {"order_date": "Not a source column", "revenue": "Amount"},
    ],
)
def test_invalid_or_ambiguous_mapping_is_rejected(client, mapping) -> None:
    content = b"When,Amount,Units,Price\n2026-01-01,10,1,10\n"

    response = _post_import(client, content, mapping)

    assert response.status_code == 422
    assert response.json()["detail"]


def test_mapping_cannot_reuse_one_numeric_source_for_date_and_revenue(client) -> None:
    content = b"Ambiguous\n45292\n"

    response = _post_import(
        client,
        content,
        {"order_date": "Ambiguous", "revenue": "Ambiguous"},
    )

    assert response.status_code == 422
    assert "more than once" in " ".join(response.json()["detail"]).lower()


def test_deeply_nested_mapping_json_returns_a_controlled_400(client) -> None:
    content = b"When,Amount\n2026-01-01,10\n"
    nested_mapping = (
        '{"order_date":'
        + "[" * 2_000
        + '"When"'
        + "]" * 2_000
        + ',"revenue":"Amount"}'
    )

    response = client.post(
        "/api/data/import",
        data={
            "mapping": nested_mapping,
            "dataset_name": "Nested mapping",
            "expected_sha256": hashlib.sha256(content).hexdigest(),
            "source_currency": "USD",
        },
        files={"file": ("sales.csv", content, "text/csv")},
    )

    assert response.status_code == 400
    detail = response.json()["detail"].lower()
    assert "mapping" in detail
    assert "source column name" in detail or "valid json object" in detail


def test_import_rejects_ambiguous_day_month_dates(client) -> None:
    content = b"When,Amount\n01/02/2026,10\n"

    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
    )

    assert response.status_code == 422
    assert "ambiguous" in " ".join(response.json()["detail"]).lower()
    assert "yyyy-mm-dd" in " ".join(response.json()["detail"]).lower()


@pytest.mark.parametrize(
    ("source_date", "message"),
    [
        ("01/02/2026 10:30", "ambiguous"),
        ("02 Jan 2026", "unsupported"),
        ("10:30:00", "unsupported"),
        ("2026-01-02T10:30:00Z", "timezone-aware"),
        ("2026-01-02 10:30+08:00", "timezone-aware"),
        ("2026", "unsupported"),
    ],
)
def test_import_rejects_ambiguous_localized_time_only_timezone_and_bare_year_dates(
    client, source_date, message
) -> None:
    content = f"When,Amount\n{source_date},10\n".encode()

    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
    )

    assert response.status_code == 422
    assert message in " ".join(response.json()["detail"]).lower()


def test_localized_date_preview_warns_that_normalization_is_required(client) -> None:
    content = b"Sale Date,Amount\n02 Jan 2026,10\n"

    response = _post_preview(client, content)

    assert response.status_code == 200
    payload = response.json()
    assert payload["suggestions"]["order_date"]["column"] == "Sale Date"
    warning_text = " ".join(payload["warnings"]).lower()
    assert "normalization before import" in warning_text
    assert "iso yyyy-mm-dd" in warning_text


@pytest.mark.parametrize(
    "source_date",
    ["2026-01-02T10:30", "2026-01-02 10:30:45"],
)
def test_import_accepts_unambiguous_local_iso_timestamp(client, source_date) -> None:
    content = f"When,Amount\n{source_date},10\n".encode()

    response = _post_import(
        client,
        content,
        {"order_date": "When", "revenue": "Amount"},
    )

    assert response.status_code == 200
    with SessionLocal() as session:
        assert session.scalar(select(SalesRecord.order_date)).isoformat() == "2026-01-02"


@pytest.mark.parametrize(
    "content",
    [
        (
            b"order_id,order_id,order_date,revenue\n"
            b"A-1,A-2,2026-01-01,10\n"
        ),
        (
            b"Order ID,order-id,order_date,revenue\n"
            b"A-1,A-2,2026-01-01,10\n"
        ),
    ],
)
def test_preview_rejects_duplicate_or_normalized_colliding_headers(client, content) -> None:
    response = _post_preview(client, content)

    assert response.status_code == 422
    assert "collide" in " ".join(response.json()["detail"]).lower()


def test_preview_rejects_ragged_csv_rows(client) -> None:
    content = b"When,Amount\n2026-01-01,10,unexpected\n"

    response = _post_preview(client, content)

    assert response.status_code == 422
    detail = " ".join(response.json()["detail"]).lower()
    assert "fields" in detail
    assert "header" in detail


def test_preview_rejects_xlsx_values_beyond_unnamed_trailing_headers(client) -> None:
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sales"
    worksheet.append(["When", "Amount", None])
    worksheet.append(["2026-01-01", 10, "unexpected"])
    workbook.save(output)
    workbook.close()

    response = _post_preview(
        client,
        output.getvalue(),
        filename="trailing-value.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    assert response.status_code == 422
    assert "beyond the last named header" in " ".join(
        response.json()["detail"]
    ).lower()


def test_lazy_xlsx_iteration_failure_returns_a_controlled_400(client, monkeypatch) -> None:
    from data_pipeline import tabular

    content = _xlsx_bytes(
        {"Sales": pd.DataFrame({"When": ["2026-01-01"], "Amount": [10]})}
    )

    class BrokenWorksheet:
        max_column = 2
        max_row = 2

        @staticmethod
        def iter_rows(*, values_only):
            assert values_only is True
            yield ("When", "Amount")
            raise ValueError("lazy worksheet corruption")

    class BrokenWorkbook:
        sheetnames = ["Sales"]

        @staticmethod
        def close():
            return None

        def __getitem__(self, sheet_name):
            assert sheet_name == "Sales"
            return BrokenWorksheet()

    monkeypatch.setattr(tabular, "load_workbook", lambda *args, **kwargs: BrokenWorkbook())

    response = _post_preview(
        client,
        content,
        filename="lazy-corruption.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    assert response.status_code == 400
    detail = response.json()["detail"].lower()
    assert "xlsx" in detail or "worksheet" in detail


@pytest.mark.parametrize(
    ("filename", "content", "content_type"),
    [
        ("sales.json", b'{"date":"2026-01-01"}', "application/json"),
        ("sales.xlsx", b"not-an-xlsx-zip", "application/octet-stream"),
        ("sales.xlsm", b"not-a-macro-workbook", "application/octet-stream"),
        ("sales.csv", b"", "text/csv"),
    ],
)
def test_preview_rejects_unsupported_or_malformed_files(
    client, filename, content, content_type
) -> None:
    response = _post_preview(
        client,
        content,
        filename=filename,
        content_type=content_type,
    )

    assert response.status_code == 400
    assert response.json()["detail"]


def test_preview_rejects_unsafe_xlsx_compression_before_workbook_parsing(client) -> None:
    archive = BytesIO()
    with zipfile.ZipFile(archive, "w", compression=zipfile.ZIP_DEFLATED) as workbook:
        workbook.writestr("[Content_Types].xml", "0" * 200_000)

    response = _post_preview(
        client,
        archive.getvalue(),
        filename="compressed.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    assert response.status_code == 400
    assert "compression ratio" in response.json()["detail"].lower()


def test_preview_enforces_workbook_column_complexity_limit(client, monkeypatch) -> None:
    from data_pipeline import tabular

    monkeypatch.setattr(tabular, "MAX_TABULAR_COLUMNS", 2)
    content = _xlsx_bytes(
        {"Sales": pd.DataFrame({"Date": ["2026-01-01"], "A": [1], "B": [2]})}
    )

    response = _post_preview(
        client,
        content,
        filename="wide.xlsx",
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    assert response.status_code == 422
    assert "column limit" in " ".join(response.json()["detail"]).lower()


def test_preview_and_import_apply_the_configured_byte_limit(client) -> None:
    original_settings = client.app.state.settings
    client.app.state.settings = replace(original_settings, max_upload_bytes=4)
    mapping = {"order_date": "When", "revenue": "Amount"}
    try:
        preview = _post_preview(client, b"12345")
        imported = _post_import(client, b"12345", mapping)
    finally:
        client.app.state.settings = original_settings

    assert preview.status_code == 413
    assert imported.status_code == 413


def test_failed_flexible_import_is_atomic_for_data_and_profile(client) -> None:
    assert client.post("/api/data/demo").status_code == 200
    before_kpis = client.get("/api/analytics/kpis").json()
    before_profile = client.get("/api/data/profile").json()
    invalid = b"When,Amount\n2026-01-01,-10\n"

    response = _post_import(
        client,
        invalid,
        {"order_date": "When", "revenue": "Amount"},
        dataset_name="Must not activate",
    )

    assert response.status_code == 422
    assert client.get("/api/analytics/kpis").json() == before_kpis
    assert client.get("/api/data/profile").json() == before_profile


def test_import_rejects_a_file_that_changed_after_preview(client) -> None:
    assert client.post("/api/data/demo").status_code == 200
    before_kpis = client.get("/api/analytics/kpis").json()
    before_profile = client.get("/api/data/profile").json()
    previewed = b"When,Amount\n2026-01-01,10\n"
    changed = b"When,Amount\n2026-01-01,999\n"

    response = _post_import(
        client,
        changed,
        {"order_date": "When", "revenue": "Amount"},
        expected_sha256=hashlib.sha256(previewed).hexdigest(),
    )

    assert response.status_code == 409
    assert "changed after preview" in response.json()["detail"].lower()
    assert client.get("/api/analytics/kpis").json() == before_kpis
    assert client.get("/api/data/profile").json() == before_profile


def test_canonical_csv_upload_remains_compatible(client) -> None:
    content = (
        b"order_id,order_date,customer_id,region,category,product,quantity,"
        b"unit_price,discount\n"
        b"A-1,2026-01-01,C-1,North,Hardware,Widget,2,10,0.1\n"
    )

    response = client.post(
        "/api/data/upload",
        files={"file": ("canonical.csv", content, "text/csv")},
    )

    assert response.status_code == 200
    assert response.json()["revenue_total"] == 18.0
    dashboard = client.get("/api/dashboard").json()
    assert dashboard["kpis"]["record_semantics"]["record_count_label"] == "Orders"
    assert client.get("/api/data/profile").status_code == 200


def test_canonical_upload_accepts_gbp_and_reloaded_profile_has_utc_timestamp(
    client,
) -> None:
    content = (
        b"order_id,order_date,customer_id,region,category,product,quantity,"
        b"unit_price,discount\n"
        b"GBP-1,2026-01-01,C-1,North,Hardware,Widget,2,10,0.1\n"
    )

    uploaded = client.post(
        "/api/data/upload",
        data={"source_currency": "GBP"},
        files={"file": ("canonical-gbp.csv", content, "text/csv")},
    )
    profile_response = client.get("/api/data/profile")

    assert uploaded.status_code == 200
    assert profile_response.status_code == 200
    profile = profile_response.json()
    assert profile["currency"] == "GBP"
    imported_at = profile["imported_at"]
    assert imported_at.endswith(("Z", "+00:00"))
    parsed = datetime.fromisoformat(imported_at.replace("Z", "+00:00"))
    assert parsed.utcoffset() == timedelta(0)
